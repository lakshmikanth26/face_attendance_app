import cv2
from tensorflow.keras.layers import DepthwiseConv2D as KerasDepthwiseConv2D
from keras.models import load_model
from PIL import Image, ImageOps
import numpy as np
import sys
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QLabel, QMessageBox
from PyQt5.QtCore import QThread, pyqtSignal
from openpyxl import Workbook, load_workbook
import os

# Disable scientific notation for clarity
np.set_printoptions(suppress=True)

class CustomDepthwiseConv2D(KerasDepthwiseConv2D):
    def __init__(self, **kwargs):
        kwargs.pop('groups', None)
        super().__init__(**kwargs)

# Load the pre-trained Keras model
model = load_model('model/keras_model.h5', compile=False, custom_objects={'DepthwiseConv2D': CustomDepthwiseConv2D})

# Load the labels
class_names = open("model/labels.txt", "r").readlines()

# Function to preprocess the image
def preprocess_frame(frame):
    size = (224, 224)
    image = Image.fromarray(frame).convert("RGB")  # Convert frame to PIL image
    image = ImageOps.fit(image, size, Image.Resampling.LANCZOS)  # Resize and crop
    image_array = np.asarray(image)  # Convert to numpy array
    normalized_image_array = (image_array.astype(np.float32) / 127.5) - 1  # Normalize
    return np.expand_dims(normalized_image_array, axis=0)  # Add batch dimension

# Initialize attendance file
def initialize_attendance_file(filename="attendance.xlsx"):
    if not os.path.exists(filename):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Status"])  # Add headers
        workbook.save(filename)

# Save attendance
def save_attendance(name, filename="attendance.xlsx"):
    workbook = load_workbook(filename)
    sheet = workbook.active
    sheet.append([name, "Present"])
    workbook.save(filename)

# Create PyQt GUI for marking attendance
class AttendanceWindow(QWidget):
    def __init__(self, name):
        super().__init__()
        self.name = name
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()

        self.label = QLabel(f"Identified student: {self.name}\nMark as Present?", self)
        layout.addWidget(self.label)

        self.yes_button = QPushButton('Yes', self)
        self.no_button = QPushButton('No', self)

        self.yes_button.clicked.connect(self.mark_attendance_yes)
        self.no_button.clicked.connect(self.mark_attendance_no)

        layout.addWidget(self.yes_button)
        layout.addWidget(self.no_button)

        self.setLayout(layout)
        self.setWindowTitle("Attendance")
        self.show()

    def mark_attendance_yes(self):
        save_attendance(self.name)
        self.close()

    def mark_attendance_no(self):
        self.close()

class ContinueScanningWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()

        self.label = QLabel("Do you want to continue scanning for another attendance?", self)
        layout.addWidget(self.label)

        self.yes_button = QPushButton('Yes', self)
        self.no_button = QPushButton('No', self)

        self.yes_button.clicked.connect(self.continue_scanning_yes)
        self.no_button.clicked.connect(self.continue_scanning_no)

        layout.addWidget(self.yes_button)
        layout.addWidget(self.no_button)

        self.setLayout(layout)
        self.setWindowTitle("Continue?")
        self.show()

    def continue_scanning_yes(self):
        self.close()
        self.continue_scan = True

    def continue_scanning_no(self):
        self.close()
        self.continue_scan = False

# Initialize webcam
cap = cv2.VideoCapture(0)

if not cap.isOpened():
    print("Error: Could not access the webcam.")
    exit()

print("Press 'q' to quit.")

initialize_attendance_file()

# Initialize QApplication
app = QApplication(sys.argv)

while True:
    ret, frame = cap.read()
    if not ret:
        print("Error: Failed to capture frame.")
        break

    # Preprocess the frame for the model
    data = preprocess_frame(frame)

    # Make predictions
    prediction = model.predict(data)
    index = np.argmax(prediction)
    class_name = class_names[index].strip()
    confidence_score = prediction[0][index]

    # Display prediction and confidence score
    cv2.putText(frame, f"Class: {class_name}, Confidence: {confidence_score:.2f}", 
                (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
    
    # Show the frame
    cv2.imshow("Webcam", frame)

    # Check confidence level
    if confidence_score >= 0.90:
        # Show attendance window
        attendance_window = AttendanceWindow(class_name)
        app.exec_()  # This is a blocking call that waits until the window is closed

        # Ask if user wants to continue scanning
        continue_window = ContinueScanningWindow()
        app.exec_()  # Wait until the user responds

        if not continue_window.continue_scan:  # If "No" clicked to stop, break the loop and close webcam
            break

    # Break loop on 'q' key press
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release the webcam and close the OpenCV window
cap.release()
cv2.destroyAllWindows()
